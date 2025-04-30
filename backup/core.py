import os
import numpy as np
import pandas as pd
from typing import List
from openpyxl.styles import Alignment
from utils import normalize_score, get_grade_level, calculate_final_score, calculate_achievement_level, adjust_column_widths

class GradeProcessor:
    def __init__(self, course_name_input, num_objectives_input, weight_inputs, usual_ratio_input, 
                 midterm_ratio_input, final_ratio_input, status_label, input_file, 
                 course_description="", objective_requirements=None):
        self.course_name_input = course_name_input
        self.num_objectives_input = num_objectives_input
        self.weight_inputs = weight_inputs
        self.usual_ratio_input = usual_ratio_input
        self.midterm_ratio_input = midterm_ratio_input
        self.final_ratio_input = final_ratio_input
        self.status_label = status_label
        self.input_file = input_file
        self.course_description = course_description
        self.objective_requirements = objective_requirements or []
        self.previous_achievement_data = None
        self.api_key = None

    def calculate_score_bounds(self, target_score: float, spread_mode: str) -> tuple:
        spread_ranges = {'large': 23, 'medium': 13, 'small': 8}
        base_spread = spread_ranges[spread_mode]

        if target_score < 40:
            spread = min(base_spread, target_score + 5)
        else:
            spread = base_spread

        min_bound = max(0.0, target_score - spread)
        max_bound = min(99.0, target_score + spread)
        
        return min_bound, max_bound

    def generate_initial_scores(self, target, n, min_bound, max_bound, dist_type):
        """生成初始整数分数，分段体现正态分布或偏态分布"""
        scores = np.zeros(n, dtype=int)
        mean = target
        std = (max_bound - min_bound) / 2  # 增大标准差，确保波动范围

        # 分段生成正态分布，增加极端值比例
        if dist_type == 'normal':
            segments = [
                (mean - 3*std, mean - 2*std, 0.10),  # 10% 低于 -2σ
                (mean - 2*std, mean - std, 0.20),    # 20% 在 -2σ 到 -1σ
                (mean - std, mean + std, 0.40),      # 40% 在 -1σ 到 +1σ
                (mean + std, mean + 2*std, 0.20),   # 20% 在 +1σ 到 +2σ
                (mean + 2*std, mean + 3*std, 0.10)  # 10% 高于 +2σ
            ]
        elif dist_type == 'left_skewed':
            segments = [
                (min_bound, mean - std, 0.1),
                (mean - std, mean, 0.2),
                (mean, mean + std, 0.4),
                (mean + std, max_bound, 0.3)
            ]
        elif dist_type == 'right_skewed':
            segments = [
                (min_bound, mean - std, 0.3),
                (mean - std, mean, 0.4),
                (mean, mean + std, 0.2),
                (mean + std, max_bound, 0.1)
            ]
        else:  # uniform
            segments = [(min_bound, max_bound, 1.0)]

        # 根据权重分配分数
        remaining_indices = list(range(n))
        for segment_min, segment_max, proportion in segments:
            num_scores = max(1, int(round(proportion * n)))  # 至少分配 1 个分数
            if num_scores == 0:
                continue
            # 确保 low < high
            low = max(int(segment_min), int(min_bound))
            high = min(int(segment_max), int(max_bound)) + 1
            if low >= high:
                low = max(int(min_bound), int(segment_min - 1))
                high = min(int(max_bound) + 1, int(segment_max + 1))
                if low >= high:
                    low = int(min_bound)
                    high = int(max_bound) + 1
            # 随机选择 num_scores 个索引
            chosen_indices = np.random.choice(remaining_indices, min(num_scores, len(remaining_indices)), replace=False)
            for idx in chosen_indices:
                try:
                    scores[idx] = np.random.randint(low, high)
                except ValueError as e:
                    print(f"Error in np.random.randint: low={low}, high={high}, error={str(e)}")
                    scores[idx] = np.random.randint(int(min_bound), int(max_bound) + 1)  # 回退到默认范围
                remaining_indices.remove(idx)

        # 确保所有分数都被赋值
        for idx in remaining_indices:
            scores[idx] = np.random.randint(int(min_bound), int(max_bound) + 1)

        return scores

    def adjust_scores(self, scores, target, weights, min_bound, max_bound, dist_type):
        """逐步调整分数以满足加权和约束，同时保留分布形态"""
        weights_array = np.array(weights)
        scores = np.array(scores, dtype=float)
        max_attempts = 1000
        attempt = 0

        while attempt < max_attempts:
            current_sum = np.sum(scores * weights_array)
            diff = target - current_sum

            if abs(diff) <= 0.1:
                break

            # 优先调整权重较大的分数
            indices = np.argsort(-weights_array)  # 按权重从大到小排序
            for idx in indices:
                weight = weights[idx]
                if weight > 0:
                    # 计算调整量（尽量整数）
                    adjustment = diff / weight
                    if adjustment > 0:
                        adjustment = max(1, int(adjustment))  # 向上调整
                    else:
                        adjustment = min(-1, int(adjustment))  # 向下调整
                    new_score = scores[idx] + adjustment
                    if min_bound <= new_score <= max_bound:
                        scores[idx] = new_score
                        break

            attempt += 1

        # 最终微调（允许小数，保留 1 位）
        current_sum = np.sum(scores * weights_array)
        diff = target - current_sum
        if abs(diff) > 0.1:
            indices = np.argsort(weights_array)  # 优先调整权重最小的分数
            for idx in indices:
                weight = weights[idx]
                if weight > 0:
                    adjustment = round(diff / weight, 1)  # 保留 1 位小数
                    new_score = scores[idx] + adjustment
                    if min_bound <= new_score <= max_bound:
                        scores[idx] = new_score
                        break

        # 再次验证加权和
        current_sum = np.sum(scores * weights_array)
        diff = target - current_sum
        if abs(diff) > 0.1:
            print(f"Warning: Final weighted sum deviation {abs(diff):.2f} exceeds 0.1 for target {target}")

        return scores

    def generate_weighted_scores(self, target_sum: float, weights: List[float], all_scores: List[List[float]], 
                                spread_mode: str = 'medium', distribution: str = 'uniform') -> List[float]:
        """
        基于分布模式和跨度范围生成成绩，确保加权和偏差 ≤ 0.1。
        """
        n = len(weights)

        # 如果 target_sum = 0，直接返回全 0 分数
        if abs(target_sum) < 0.0001:
            return np.zeros(n).tolist()

        # 计算分数范围
        min_bound, max_bound = self.calculate_score_bounds(target_sum, spread_mode)

        # 生成初始整数分数
        scores = self.generate_initial_scores(target_sum, n, min_bound, max_bound, distribution)

        # 逐步调整分数以满足加权和约束
        optimized_scores = self.adjust_scores(scores, target_sum, weights, min_bound, max_bound, distribution)

        # 调试日志
        print(f"Generated scores: {optimized_scores.tolist()}")
        print(f"Distribution - Mean: {np.mean(optimized_scores):.2f}, Std: {np.std(optimized_scores):.2f}")

        return optimized_scores.tolist()

    def process_grades(self, num_objectives, weights, usual_ratio, midterm_ratio, final_ratio, 
                      spread_mode='medium', distribution='uniform'):
        """处理成绩数据"""
        course_name = self.course_name_input.text()
        if not course_name:
            raise ValueError("请输入课程名称")
            
        output_dir = os.path.dirname(self.input_file)
        detail_output = os.path.join(output_dir, f'{course_name}成绩单详情.xlsx')

        # 读取 Excel 文件并校验列名
        try:
            df = pd.read_excel(self.input_file)
        except Exception as e:
            raise ValueError(f"无法读取输入文件: {str(e)}")

        required_columns = ['学生姓名', '平时成绩', '期中成绩', '期末成绩', '总和']
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            raise ValueError(f"输入文件缺少以下必需列：{', '.join(missing_columns)}。请确保文件包含：{', '.join(required_columns)}")
        
        result_data = []
        all_usual_scores = [[] for _ in range(num_objectives)]
        all_midterm_scores = [[] for _ in range(num_objectives)]
        all_final_scores = [[] for _ in range(num_objectives)]
        
        for idx, row in df.iterrows():
            self.status_label.setText(f"正在处理第 {idx+1}/{len(df)} 个学生的成绩...")
            name = row['学生姓名']
            total_usual = row['平时成绩']
            total_midterm = row['期中成绩']
            total_final = row['期末成绩']
            total_score = row['总和']
            
            try:
                usual_scores = self.generate_weighted_scores(total_usual, weights, all_usual_scores, spread_mode, distribution)
                midterm_scores = self.generate_weighted_scores(total_midterm, weights, all_midterm_scores, spread_mode, distribution)
                final_scores = self.generate_weighted_scores(total_final, weights, all_final_scores, spread_mode, distribution)
            except Exception as e:
                print(f"Error generating scores for student {name}: {str(e)}")
                raise

            for i in range(num_objectives):
                all_usual_scores[i].append(usual_scores[i])
                all_midterm_scores[i].append(midterm_scores[i])
                all_final_scores[i].append(final_scores[i])
            
            for i in range(num_objectives):
                score = calculate_final_score(
                    usual_scores[i], midterm_scores[i], final_scores[i],
                    usual_ratio, midterm_ratio, final_ratio
                )
                
                result_data.append({
                    '学生姓名': name,
                    '课程目标': i + 1,
                    '平时成绩': usual_scores[i],
                    '期中成绩': midterm_scores[i],
                    '期末成绩': final_scores[i],
                    '权重': weights[i],
                    '平时成绩占比': usual_ratio,
                    '期中成绩占比': midterm_ratio,
                    '期末成绩占比': final_ratio,
                    '分数': score,
                    '等级': get_grade_level(score)
                })
            
            final_total_score = calculate_final_score(
                total_usual, total_midterm, total_final,
                usual_ratio, midterm_ratio, final_ratio
            )
            
            result_data.append({
                '学生姓名': name,
                '课程目标': '总和',
                '平时成绩': total_usual,
                '期中成绩': total_midterm,
                '期末成绩': total_final,
                '权重': sum(weights),
                '平时成绩占比': usual_ratio,
                '期中成绩占比': midterm_ratio,
                '期末成绩占比': final_ratio,
                '分数': final_total_score,
                '等级': get_grade_level(final_total_score)
            })
        
        result_df = pd.DataFrame(result_data)
        
        try:
            with pd.ExcelWriter(detail_output, engine='openpyxl') as writer:
                result_df.to_excel(writer, index=False, sheet_name='Sheet1')
                
                worksheet = writer.sheets['Sheet1']
                if self.course_description:
                    worksheet.cell(row=1, column=1).value = f"课程简介: {self.course_description}"
                
                current_name = None
                start_row = 2 if not self.course_description else 3
                
                for i, name in enumerate(result_df['学生姓名'], start=start_row):
                    if name != current_name:
                        if current_name is not None and start_row != i-1:
                            worksheet.merge_cells(f'A{start_row}:A{i-1}')
                        current_name = name
                        start_row = i
                
                if start_row != i:
                    worksheet.merge_cells(f'A{start_row}:A{i}')
                    
                adjust_column_widths(worksheet)
        except Exception as e:
            print(f"Error writing to Excel: {str(e)}")
            raise

        overall_achievement = self.generate_objective_analysis_report(result_df, course_name, weights, usual_ratio, midterm_ratio, final_ratio)
        return overall_achievement

    def generate_objective_analysis_report(self, result_df: pd.DataFrame, course_name: str, weights, usual_ratio, midterm_ratio, final_ratio) -> float:
        """生成课程目标达成度分析报告"""
        output_dir = os.path.dirname(self.input_file)
        analysis_output = os.path.join(output_dir, f'{course_name}课程目标达成度分析表.xlsx')
        
        objectives = sorted([i for i in result_df['课程目标'].unique() if isinstance(i, int)])
        
        analysis_data = []
        
        exam_types = [
            ('平时考核\n(A)', '平时成绩'),
            ('期中考核\n(B)', '期中成绩'),
            ('期末考核\n(C)', '期末成绩')
        ]
        
        weights_dict = {f'课程目标{obj}': round(w * 100, 3) for obj, w in zip(objectives, weights)}
        
        m_values = {}
        
        for exam_name, score_column in exam_types:
            avg_scores = {}
            score_ratios = {}
            
            for obj in objectives:
                obj_scores = result_df[
                    (result_df['课程目标'] == obj)
                ][score_column].tolist()
                
                if obj_scores:
                    avg_scores[f'课程目标{obj}'] = round(np.mean(obj_scores), 1)
                    score_ratios[f'课程目标{obj}'] = 100
            
            analysis_data.extend([
                {
                    '考核环节': exam_name,
                    '指标类型': '平均分',
                    **avg_scores
                },
                {
                    '考核环节': exam_name,
                    '指标类型': '分值/满分\n(S)',
                    **score_ratios
                },
                {
                    '考核环节': exam_name,
                    '指标类型': '分权重 (K)',
                    **weights_dict
                }
            ])
        
        m_row = {'考核环节': '课程分目标达成度\n(M)', '指标类型': ''}
        for obj in objectives:
            usual_avg = analysis_data[0].get(f'课程目标{obj}', 0)
            midterm_avg = analysis_data[3].get(f'课程目标{obj}', 0)
            final_avg = analysis_data[6].get(f'课程目标{obj}', 0)
            # 修正 M 值计算公式
            m = usual_avg * usual_ratio + midterm_avg * midterm_ratio + final_avg * final_ratio
            m_row[f'课程目标{obj}'] = round(m, 1)
            m_values[obj] = m
        
        analysis_data.append(m_row)
        
        z_row = {'考核环节': '课程分目标总权重\n(Z)', '指标类型': ''}
        for obj in objectives:
            z_row[f'课程目标{obj}'] = weights_dict[f'课程目标{obj}']
        analysis_data.append(z_row)
        
        total_achievement = sum(m_values[obj] * weights[obj-1] for obj in objectives)
        total_achievement = round(total_achievement, 1)
        total_row = {'考核环节': '课程总目标达成度', '指标类型': ''}
        for obj in objectives:
            total_row[f'课程目标{obj}'] = total_achievement if obj == objectives[0] else ''
        analysis_data.append(total_row)
        
        columns = ['考核环节', '指标类型'] + [f'课程目标{i}' for i in objectives]
        analysis_df = pd.DataFrame(analysis_data, columns=columns)
        
        try:
            with pd.ExcelWriter(analysis_output, engine='openpyxl') as writer:
                analysis_df.to_excel(writer, index=False, sheet_name='Sheet1')
                
                worksheet = writer.sheets['Sheet1']
                
                current_exam = None
                start_row = 2
                for i, exam in enumerate(analysis_df['考核环节'], start=2):
                    if exam != current_exam:
                        if current_exam is not None and start_row != i:
                            worksheet.merge_cells(f'A{start_row}:A{i-1}')
                        current_exam = exam
                        start_row = i
                
                if start_row != i:
                    worksheet.merge_cells(f'A{start_row}:A{i}')
                
                m_row_idx = len(analysis_df) - 2 + 1
                worksheet.merge_cells(f'A{m_row_idx}:B{m_row_idx}')
                
                z_row_idx = len(analysis_df) - 1 + 1
                worksheet.merge_cells(f'A{z_row_idx}:B{z_row_idx}')
                
                total_row_idx = len(analysis_df) + 1
                worksheet.merge_cells(f'A{total_row_idx}:B{total_row_idx}')
                worksheet.merge_cells(f'C{total_row_idx}:{chr(ord("C") + len(objectives) - 1)}{total_row_idx}')
                worksheet[f'C{total_row_idx}'].alignment = Alignment(horizontal='center', vertical='center')
                
                adjust_column_widths(worksheet)
        except Exception as e:
            print(f"Error writing to Excel: {str(e)}")
            raise

        return total_achievement

    def load_previous_achievement(self, file_path: str) -> None:
        """加载上一学年达成度表"""
        if not file_path:
            return
        try:
            self.previous_achievement_data = pd.read_excel(file_path)
            if self.status_label:
                self.status_label.setText(f"已加载上一学年达成度表: {os.path.basename(file_path)}")
        except Exception as e:
            if self.status_label:
                self.status_label.setText("加载上一学年达成度表失败！")
            raise ValueError(f"加载上一学年达成度表失败: {str(e)}")

    def store_api_key(self, api_key: str) -> None:
        """存储API Key"""
        self.api_key = api_key
        if self.status_label:
            self.status_label.setText("已存储API Key")

    def generate_ai_report(self) -> None:
        """生成AI分析报告（占位）"""
        if not self.api_key:
            if self.status_label:
                self.status_label.setText("请先设置API Key")
            return
        if self.status_label:
            self.status_label.setText(f"API Key: {self.api_key}，AI分析报告功能待实现")