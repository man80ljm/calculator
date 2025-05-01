import os
import numpy as np
import pandas as pd
import json
import requests
from typing import List, Dict,Callable, Optional
from openpyxl.styles import Alignment, PatternFill, Font
import openpyxl
from utils import normalize_score, get_grade_level, calculate_final_score, calculate_achievement_level, adjust_column_widths
import time

class GradeProcessor:
    def __init__(self, course_name_input, num_objectives_input, weight_inputs, usual_ratio_input, 
                 midterm_ratio_input, final_ratio_input, status_label, input_file, 
                 course_description="", objective_requirements=None):
        self.course_name_input = course_name_input
        self.num_objectives_input = num_objectives_input
        self.weight_inputs = weight_inputs
        self.usual_ratio_input = usual_ratio_input
        self.midterm_ratio_input = midterm_ratio_input
        self.final_ratio_input = final_ratio_input
        self.status_label = status_label
        self.input_file = input_file
        self.course_description = course_description
        self.objective_requirements = objective_requirements or []
        self.previous_achievement_data = None
        self.api_key = None

    def test_deepseek_api(self, api_key: str) -> str:
        """测试 DeepSeek API 连接"""
        url = "https://api.deepseek.com/v1/chat/completions"
        api_key = api_key.strip().strip('<').strip('>')
        headers = {
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json"
        }
        payload = {
            "model": "deepseek-chat",
            "messages": [
                {"role": "system", "content": "You are a helpful assistant."},
                {"role": "user", "content": "测试连接"}
            ],
            "temperature": 0.7,
            "top_p": 1,
            "max_tokens": 10,
            "stream": False
        }
        
        try:
            response = requests.post(url, headers=headers, json=payload, timeout=10)
            response.raise_for_status()
            return "连接成功"
        except requests.RequestException as e:
            error_message = f"连接失败: {str(e)}"
            if hasattr(e, 'response') and e.response is not None:
                error_message += f"\n服务器返回: {e.response.text}"
            return error_message

    def call_deepseek_api(self, prompt: str) -> str:
        """调用 DeepSeek API 获取答案，增加重试机制"""
        if not self.api_key:
            return "请先设置API Key"
        
        url = "https://api.deepseek.com/v1/chat/completions"
        api_key = self.api_key.strip().strip('<').strip('>')
        headers = {
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json"
        }
        # 根据问题类型设置不同的字数限制
        if "针对上一年度存在问题的改进情况" in prompt:
            max_tokens = 200  # 200 字以内
            prompt = f"{prompt}\n用一段话回答，不要分点阐述，同时控制字数在200字以内。"
        else:
            max_tokens = 100  # 100 字以内
            prompt = f"{prompt}\n用一段话回答，不要分点阐述，同时控制字数在100字以内。"
        
        payload = {
            "model": "deepseek-chat",
            "messages": [
                {"role": "system", "content": "You are a helpful assistant specializing in course analysis and improvement."},
                {"role": "user", "content": prompt}
            ],
            "temperature": 0.7,
            "top_p": 1,
            "max_tokens": max_tokens,
            "stream": False
        }
        
        max_retries = 3
        for attempt in range(max_retries):
            try:
                response = requests.post(url, headers=headers, json=payload, timeout=30)
                response.raise_for_status()
                return response.json()['choices'][0]['message']['content'].strip()
            except requests.Timeout:
                if attempt < max_retries - 1:
                    print(f"API 调用超时，正在重试（第 {attempt + 1}/{max_retries} 次）...")
                    time.sleep(2)
                    continue
                return "API 调用超时，请检查网络连接或稍后重试（可能需要使用 VPN 或代理访问 api.deepseek.com）"
            except requests.RequestException as e:
                error_message = f"API 调用失败: {str(e)}"
                if hasattr(e, 'response') and e.response is not None:
                    error_message += f"\n服务器返回: {e.response.text}"
                if attempt < max_retries - 1:
                    print(f"API 调用失败，正在重试（第 {attempt + 1}/{max_retries} 次）...")
                    time.sleep(2)
                    continue
                return error_message
            except (KeyError, IndexError):
                return "API 返回格式错误，无法解析结果"

    def calculate_score_bounds(self, target_score: float, spread_mode: str) -> tuple:
        spread_ranges = {'large': 23, 'medium': 13, 'small': 8}
        base_spread = spread_ranges[spread_mode]

        if target_score < 40:
            spread = min(base_spread, target_score + 5)
        else:
            spread = base_spread

        min_bound = max(0.0, target_score - spread)
        max_bound = min(99.0, target_score + spread)
        
        return min_bound, max_bound

    def generate_initial_scores(self, target, n, min_bound, max_bound, dist_type):
        """生成初始整数分数，分段体现正态分布或偏态分布"""
        scores = np.zeros(n, dtype=int)
        mean = target
        std = (max_bound - min_bound) / 2

        if dist_type == 'normal':
            segments = [
                (mean - 3*std, mean - 2*std, 0.10),
                (mean - 2*std, mean - std, 0.20),
                (mean - std, mean + std, 0.40),
                (mean + std, mean + 2*std, 0.20),
                (mean + 2*std, mean + 3*std, 0.10)
            ]
        elif dist_type == 'left_skewed':
            segments = [
                (min_bound, mean - std, 0.1),
                (mean - std, mean, 0.2),
                (mean, mean + std, 0.4),
                (mean + std, max_bound, 0.3)
            ]
        elif dist_type == 'right_skewed':
            segments = [
                (min_bound, mean - std, 0.3),
                (mean - std, mean, 0.4),
                (mean, mean + std, 0.2),
                (mean + std, max_bound, 0.1)
            ]
        else:
            segments = [(min_bound, max_bound, 1.0)]

        remaining_indices = list(range(n))
        for segment_min, segment_max, proportion in segments:
            num_scores = max(1, int(round(proportion * n)))
            if num_scores == 0:
                continue
            low = max(int(segment_min), int(min_bound))
            high = min(int(segment_max), int(max_bound)) + 1
            if low >= high:
                low = max(int(min_bound), int(segment_min - 1))
                high = min(int(max_bound) + 1, int(segment_max + 1))
                if low >= high:
                    low = int(min_bound)
                    high = int(max_bound) + 1
            chosen_indices = np.random.choice(remaining_indices, min(num_scores, len(remaining_indices)), replace=False)
            for idx in chosen_indices:
                try:
                    scores[idx] = np.random.randint(low, high)
                except ValueError as e:
                    print(f"Error in np.random.randint: low={low}, high={high}, error={str(e)}")
                    scores[idx] = np.random.randint(int(min_bound), int(max_bound) + 1)
                remaining_indices.remove(idx)

        for idx in remaining_indices:
            scores[idx] = np.random.randint(int(min_bound), int(max_bound) + 1)

        return scores

    def adjust_scores(self, scores, target, weights, min_bound, max_bound, dist_type):
        """逐步调整分数以满足加权和约束，同时保留分布形态"""
        weights_array = np.array(weights)
        scores = np.array(scores, dtype=float)
        max_attempts = 1000
        attempt = 0

        while attempt < max_attempts:
            current_sum = np.sum(scores * weights_array)
            diff = target - current_sum

            if abs(diff) <= 0.1:
                break

            indices = np.argsort(-weights_array)
            for idx in indices:
                weight = weights[idx]
                if weight > 0:
                    adjustment = diff / weight
                    if adjustment > 0:
                        adjustment = max(1, int(adjustment))
                    else:
                        adjustment = min(-1, int(adjustment))
                    new_score = scores[idx] + adjustment
                    if min_bound <= new_score <= max_bound:
                        scores[idx] = new_score
                        break

            attempt += 1

        current_sum = np.sum(scores * weights_array)
        diff = target - current_sum
        if abs(diff) > 0.1:
            indices = np.argsort(weights_array)
            for idx in indices:
                weight = weights[idx]
                if weight > 0:
                    adjustment = round(diff / weight, 1)
                    new_score = scores[idx] + adjustment
                    if min_bound <= new_score <= max_bound:
                        scores[idx] = new_score
                        break

        current_sum = np.sum(scores * weights_array)
        diff = target - current_sum
        if abs(diff) > 0.1:
            print(f"Warning: Final weighted sum deviation {abs(diff):.2f} exceeds 0.1 for target {target}")

        return scores

    def generate_weighted_scores(self, target_sum: float, weights: List[float], all_scores: List[List[float]], 
                                spread_mode: str = 'medium', distribution: str = 'uniform') -> List[float]:
        """
        基于分布模式和跨度范围生成成绩，确保加权和偏差 ≤ 0.1。
        """
        n = len(weights)

        if abs(target_sum) < 0.0001:
            return np.zeros(n).tolist()

        min_bound, max_bound = self.calculate_score_bounds(target_sum, spread_mode)
        scores = self.generate_initial_scores(target_sum, n, min_bound, max_bound, distribution)
        optimized_scores = self.adjust_scores(scores, target_sum, weights, min_bound, max_bound, distribution)

        print(f"Generated scores: {optimized_scores.tolist()}")
        print(f"Distribution - Mean: {np.mean(optimized_scores):.2f}, Std: {np.std(optimized_scores):.2f}")

        return optimized_scores.tolist()

    def process_grades(self, num_objectives, weights, usual_ratio, midterm_ratio, final_ratio, 
                      spread_mode='medium', distribution='uniform',progress_callback: Optional[Callable[[int], None]] = None):
        """处理成绩数据"""
        course_name = self.course_name_input.text()
        if not course_name:
            raise ValueError("请输入课程名称")
            
        output_dir = os.path.dirname(self.input_file)
        detail_output = os.path.join(output_dir, f'{course_name}成绩单详情.xlsx')

        try:
            df = pd.read_excel(self.input_file)
        except Exception as e:
            raise ValueError(f"无法读取输入文件: {str(e)}")

        required_columns = ['学生姓名', '平时成绩', '期中成绩', '期末成绩', '总和']
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            raise ValueError(f"输入文件缺少以下必需列：{', '.join(missing_columns)}。请确保文件包含：{', '.join(required_columns)}")
        
        result_data = []
        all_usual_scores = [[] for _ in range(num_objectives)]
        all_midterm_scores = [[] for _ in range(num_objectives)]
        all_final_scores = [[] for _ in range(num_objectives)]
        
        for idx, row in df.iterrows():
            self.status_label.setText(f"正在处理第 {idx+1}/{len(df)} 个学生的成绩...")
            if progress_callback:
                progress_callback(idx)  # 调用进度回调
            name = row['学生姓名']
            total_usual = row['平时成绩']
            total_midterm = row['期中成绩']
            total_final = row['期末成绩']
            total_score = row['总和']
            
            try:
                usual_scores = self.generate_weighted_scores(total_usual, weights, all_usual_scores, spread_mode, distribution)
                midterm_scores = self.generate_weighted_scores(total_midterm, weights, all_midterm_scores, spread_mode, distribution)
                final_scores = self.generate_weighted_scores(total_final, weights, all_final_scores, spread_mode, distribution)
            except Exception as e:
                print(f"Error generating scores for student {name}: {str(e)}")
                raise

            for i in range(num_objectives):
                all_usual_scores[i].append(usual_scores[i])
                all_midterm_scores[i].append(midterm_scores[i])
                all_final_scores[i].append(final_scores[i])
            
            for i in range(num_objectives):
                score = calculate_final_score(
                    usual_scores[i], midterm_scores[i], final_scores[i],
                    usual_ratio, midterm_ratio, final_ratio
                )
                
                result_data.append({
                    '学生姓名': name,
                    '课程目标': i + 1,
                    '平时成绩': usual_scores[i],
                    '期中成绩': midterm_scores[i],
                    '期末成绩': final_scores[i],
                    '权重': weights[i],
                    '平时成绩占比': usual_ratio,
                    '期中成绩占比': midterm_ratio,
                    '期末成绩占比': final_ratio,
                    '分数': score,
                    '等级': get_grade_level(score)
                })
            
            final_total_score = calculate_final_score(
                total_usual, total_midterm, total_final,
                usual_ratio, midterm_ratio, final_ratio
            )
            
            result_data.append({
                '学生姓名': name,
                '课程目标': '总和',
                '平时成绩': total_usual,
                '期中成绩': total_midterm,
                '期末成绩': total_final,
                '权重': sum(weights),
                '平时成绩占比': usual_ratio,
                '期中成绩占比': midterm_ratio,
                '期末成绩占比': final_ratio,
                '分数': final_total_score,
                '等级': get_grade_level(final_total_score)
            })
        
        result_df = pd.DataFrame(result_data)
        
        try:
            with pd.ExcelWriter(detail_output, engine='openpyxl') as writer:
                # 表格从第 1 行开始写入（startrow=0），移除课程简介
                start_row = 0
                result_df.to_excel(writer, index=False, sheet_name='Sheet1', startrow=start_row)
                
                worksheet = writer.sheets['Sheet1']
                
                # 设置表格内容居中（包括表头）
                table_start_row = start_row + 1
                table_end_row = table_start_row + len(result_df)
                num_columns = len(result_df.columns)
                for row in range(table_start_row, table_end_row + 1):
                    for col in range(1, num_columns + 1):
                        cell = worksheet.cell(row=row, column=col)
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # 标记“总和”行为黄色
                yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                for row_idx, row_data in result_df.iterrows():
                    if row_data['课程目标'] == '总和':
                        excel_row = table_start_row + row_idx + 1  # 转换为 Excel 行号
                        for col in range(1, num_columns + 1):
                            cell = worksheet.cell(row=excel_row, column=col)
                            cell.fill = yellow_fill
                
                # 先调整列宽，再进行合并操作
                adjust_column_widths(worksheet)
                
                # 强制设置姓名列（A 列）宽度为 8 个字符
                worksheet.column_dimensions['A'].width = 8
                
                # 合并姓名列相同的单元格
                current_name = None
                merge_start_row = table_start_row + 1  # 跳过表头行
                for i, name in enumerate(result_df['学生姓名'], start=merge_start_row):
                    if name != current_name:
                        if current_name is not None and merge_start_row != i:
                            worksheet.merge_cells(f'A{merge_start_row}:A{i-1}')
                        current_name = name
                        merge_start_row = i
                
                if merge_start_row != i:
                    worksheet.merge_cells(f'A{merge_start_row}:A{i}')
        except Exception as e:
            print(f"Error writing to Excel: {str(e)}")
            raise

        overall_achievement = self.generate_objective_analysis_report(result_df, course_name, weights, usual_ratio, midterm_ratio, final_ratio)
        return overall_achievement

    def generate_objective_analysis_report(self, result_df: pd.DataFrame, course_name: str, weights, usual_ratio, midterm_ratio, final_ratio) -> float:
        """生成课程目标达成度分析报告"""
        output_dir = os.path.dirname(self.input_file)
        analysis_output = os.path.join(output_dir, f'{course_name}课程目标达成度分析表.xlsx')
        
        objectives = sorted([i for i in result_df['课程目标'].unique() if isinstance(i, int)])
        
        analysis_data = []
        
        exam_types = [
            ('平时考核\n(A)', '平时成绩'),
            ('期中考核\n(B)', '期中成绩'),
            ('期末考核\n(C)', '期末成绩')
        ]
        
        weights_dict = {f'课程目标{obj}': round(w * 100, 3) for obj, w in zip(objectives, weights)}
        
        m_values = {}
        
        for exam_name, score_column in exam_types:
            avg_scores = {}
            score_ratios = {}
            
            for obj in objectives:
                obj_scores = result_df[
                    (result_df['课程目标'] == obj)
                ][score_column].tolist()
                
                if obj_scores:
                    avg_scores[f'课程目标{obj}'] = round(np.mean(obj_scores), 1)
                    score_ratios[f'课程目标{obj}'] = 100
            
            analysis_data.extend([
                {
                    '考核环节': exam_name,
                    '指标类型': '平均分',
                    **avg_scores
                },
                {
                    '考核环节': exam_name,
                    '指标类型': '分值/满分\n(S)',
                    **score_ratios
                },
                {
                    '考核环节': exam_name,
                    '指标类型': '分权重 (K)',
                    **weights_dict
                }
            ])
        
        m_row = {'考核环节': '课程分目标达成度\n(M)', '指标类型': ''}
        for obj in objectives:
            usual_avg = analysis_data[0].get(f'课程目标{obj}', 0)
            midterm_avg = analysis_data[3].get(f'课程目标{obj}', 0)
            final_avg = analysis_data[6].get(f'课程目标{obj}', 0)
            m = usual_avg * usual_ratio + midterm_avg * midterm_ratio + final_avg * final_ratio
            m_row[f'课程目标{obj}'] = round(m, 1)
            m_values[obj] = m
        
        analysis_data.append(m_row)
        
        z_row = {'考核环节': '课程分目标总权重\n(Z)', '指标类型': ''}
        for obj in objectives:
            z_row[f'课程目标{obj}'] = weights_dict[f'课程目标{obj}']
        analysis_data.append(z_row)
        
        total_achievement = sum(m_values[obj] * weights[obj-1] for obj in objectives)
        total_achievement = round(total_achievement, 1)
        total_row = {'考核环节': '课程总目标达成度', '指标类型': ''}
        for obj in objectives:
            total_row[f'课程目标{obj}'] = total_achievement if obj == objectives[0] else ''
        analysis_data.append(total_row)
        
        columns = ['考核环节', '指标类型'] + [f'课程目标{i}' for i in objectives]
        analysis_df = pd.DataFrame(analysis_data, columns=columns)
        
        try:
            with pd.ExcelWriter(analysis_output, engine='openpyxl') as writer:
                analysis_df.to_excel(writer, index=False, sheet_name='Sheet1')
                
                worksheet = writer.sheets['Sheet1']
                
                # 设置“课程目标”列（从 C 列开始）内容居中
                num_rows = len(analysis_df) + 1  # 包括表头
                for col in range(3, 3 + len(objectives)):  # C 列到 C+len(objectives)-1 列
                    for row in range(1, num_rows + 1):
                        cell = worksheet.cell(row=row, column=col)
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                
                current_exam = None
                start_row = 2
                for i, exam in enumerate(analysis_df['考核环节'], start=2):
                    if exam != current_exam:
                        if current_exam is not None and start_row != i:
                            worksheet.merge_cells(f'A{start_row}:A{i-1}')
                        current_exam = exam
                        start_row = i
                
                if start_row != i:
                    worksheet.merge_cells(f'A{start_row}:A{i}')
                
                m_row_idx = len(analysis_df) - 2 + 1
                worksheet.merge_cells(f'A{m_row_idx}:B{m_row_idx}')
                
                z_row_idx = len(analysis_df) - 1 + 1
                worksheet.merge_cells(f'A{z_row_idx}:B{z_row_idx}')
                
                total_row_idx = len(analysis_df) + 1
                worksheet.merge_cells(f'A{total_row_idx}:B{total_row_idx}')
                worksheet.merge_cells(f'C{total_row_idx}:{chr(ord("C") + len(objectives) - 1)}{total_row_idx}')
                worksheet[f'C{total_row_idx}'].alignment = Alignment(horizontal='center', vertical='center')
                
                adjust_column_widths(worksheet)
        except Exception as e:
            print(f"Error writing to Excel: {str(e)}")
            raise

        return total_achievement

    def load_previous_achievement(self, file_path: str) -> None:
        """加载上一学年达成度表，处理目标数量不一致的情况"""
        if not file_path:
            # 如果没有文件，初始化默认值（全部为 0）
            self.previous_achievement_data = {f'课程目标{i}': 0 for i in range(1, 6)}
            self.previous_achievement_data['课程总目标'] = 0
            return
        
        try:
            # 检查文件是否存在
            if not os.path.exists(file_path):
                # 如果文件不存在，初始化默认值，而不是抛出异常
                self.previous_achievement_data = {f'课程目标{i}': 0 for i in range(1, 6)}
                self.previous_achievement_data['课程总目标'] = 0
                if self.status_label:
                    self.status_label.setText("未找到上一学年达成度表，已使用默认值")
                return
            
            df = pd.read_excel(file_path)
            print(f"加载文件: {file_path}")
            print(f"表格列名: {df.columns.tolist()}")
            
            if '考核环节' in df.columns:
                print("检测到程序生成的达成度分析表，尝试解析...")
                data = {f'课程目标{i}': 0 for i in range(1, 6)}
                data['课程总目标'] = 0
                
                m_row = df[df['考核环节'] == '课程分目标达成度\n(M)']
                if not m_row.empty:
                    print(f"找到'课程分目标达成度(M)'行: {m_row.to_dict()}")
                    for i in range(1, 6):
                        col_name = f'课程目标{i}'
                        if col_name in m_row.columns and pd.notna(m_row[col_name].iloc[0]):
                            data[col_name] = float(m_row[col_name].iloc[0])
                            print(f"提取 {col_name}: {data[col_name]}")
                
                total_row = df[df['考核环节'] == '课程总目标达成度']
                if not total_row.empty:
                    print(f"找到'课程总目标达成度'行: {total_row.to_dict()}")
                    for col in total_row.columns:
                        if col.startswith('课程目标') and pd.notna(total_row[col].iloc[0]):
                            data['课程总目标'] = float(total_row[col].iloc[0])
                            print(f"提取 课程总目标: {data['课程总目标']}")
                            break
                
                self.previous_achievement_data = data
            else:
                print("未检测到'考核环节'列，尝试按照简单格式解析...")
                required_columns = ['课程目标', '上一年度达成度']
                missing_columns = [col for col in required_columns if col not in df.columns]
                if missing_columns:
                    raise ValueError(f"上一学年达成度表缺少以下必需列：{', '.join(missing_columns)}。请确保文件包含：{', '.join(required_columns)}。当前列名：{', '.join(df.columns)}")
                
                data = {f'课程目标{i}': 0 for i in range(1, 6)}
                data['课程总目标'] = 0
                
                for _, row in df.iterrows():
                    target = str(row['课程目标']).strip()
                    if target in data and isinstance(row['上一年度达成度'], (int, float)):
                        data[target] = float(row['上一年度达成度'])
                
                self.previous_achievement_data = data
            
            if self.status_label:
                self.status_label.setText(f"已加载上一学年达成度表: {os.path.basename(file_path)}")
        except Exception as e:
            if self.status_label:
                self.status_label.setText("加载上一学年达成度表失败！")
            raise ValueError(f"加载上一学年达成度表失败: {str(e)}")

    def generate_improvement_report(self, current_achievement: Dict[str, float], course_name: str, num_objectives: int, answers=None) -> None:
        """生成课程持续改进机制信息报告"""
        output_dir = os.path.dirname(self.input_file)
        output_file = os.path.join(output_dir, f'{course_name}持续改进机制信息.xlsx')
        
        df_data = []
        for i in range(1, 6):
            prev_score = self.previous_achievement_data.get(f'课程目标{i}', 0)
            current_score = current_achievement.get(f'课程目标{i}', 0)
            next_score = current_score + 2 if current_score > 0 else 0
            row = {
                '课程目标': f'课程目标{i}',
                '上一年度达成度': prev_score,
                '本一年度目标达成度': current_score,
                '本次达程度': 0,
                '下一年度目标达程度': next_score
            }
            df_data.append(row)
        
        prev_total = self.previous_achievement_data.get('课程总目标', 0)
        current_total = current_achievement.get('总达成度', 0)
        next_total = current_total + 2 if current_total > 0 else 0
        df_data.append({
            '课程目标': '课程总目标',
            '上一年度达成度': prev_total,
            '本一年度目标达成度': current_total,
            '本次达程度': 0,
            '下一年度目标达程度': next_total
        })
        
        questions = ["针对上一年度存在问题的改进情况"]
        for i in range(1, 6):
            questions.append(f"课程目标{i}达成情况分析")
            questions.append(f"该课程目标{i}达成情况存在问题分析及改进措施")
        
        # 使用传入的 answers 或生成新答案
        if answers is None:
            context = f"课程简介: {self.course_description}\n"
            for i, req in enumerate(self.objective_requirements, 1):
                context += f"课程目标{i}要求: {req}\n"
            for i in range(1, 6):
                prev_score = self.previous_achievement_data.get(f'课程目标{i}', 0)
                current_score = current_achievement.get(f'课程目标{i}', 0)
                context += f"课程目标{i}上一年度达成度: {prev_score}\n"
                context += f"课程目标{i}本年度达成度: {current_score}\n"
            context += f"课程总目标上一年度达成度: {prev_total}\n"
            context += f"课程总目标本年度达成度: {current_total}\n"
            
            cache_file = os.path.join(output_dir, 'api_cache.json')
            cached_answers = {}
            if os.path.exists(cache_file):
                try:
                    with open(cache_file, 'r', encoding='utf-8') as f:
                        cached_answers = json.load(f)
                except Exception as e:
                    print(f"加载缓存失败: {str(e)}")
            
            answers = []
            total_questions = len(questions)
            for i, question in enumerate(questions):
                if self.status_label:
                    self.status_label.setText(f"正在处理第 {i+1}/{total_questions} 个问题...")
                if "课程目标" in question and int(question.split('课程目标')[1][0]) > num_objectives:
                    answers.append("无")
                    continue
                prompt = f"{context}\n问题: {question}"
                cache_key = f"{course_name}_{question}"
                if cache_key in cached_answers:
                    answers.append(cached_answers[cache_key])
                else:
                    answer = self.call_deepseek_api(prompt)
                    cached_answers[cache_key] = answer
                    answers.append(answer)
                    try:
                        with open(cache_file, 'w', encoding='utf-8') as f:
                            json.dump(cached_answers, f, indent=4, ensure_ascii=False)
                    except Exception as e:
                        print(f"保存缓存失败: {str(e)}")
        
        df = pd.DataFrame(df_data)
        try:
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Sheet1', startrow=0)
                
                worksheet = writer.sheets['Sheet1']
                
                # 设置“课程目标1”到“课程总目标”行的内容居中（前 6 行）
                for row in range(2, 8):
                    for col in range(1, 6):
                        cell = worksheet.cell(row=row, column=col)
                        cell.alignment = Alignment(horizontal='center', vertical='center')

                # 写入 DeepSeek API 的问题和回答
                start_row = len(df) + 2  # 第 8 行开始
                # 添加分类标题“课程目标达成情况、存在问题分析及改进措施”
                worksheet[f'A{start_row}'].value = "课程目标达成情况、存在问题分析及改进措施"
                worksheet[f'A{start_row}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                end_row = start_row + len(questions) - 1  # 第 18 行结束
                worksheet.merge_cells(f'A{start_row}:A{end_row}')

                # 写入问题和回答，并设置格式
                for i, (question, answer) in enumerate(zip(questions, answers)):
                    row = start_row + i
                    worksheet[f'B{row}'].value = question
                    worksheet[f'C{row}'].value = answer
                    worksheet.merge_cells(f'C{row}:E{row}')
                    # 设置字体大小为 10 号，垂直居中
                    cell_b = worksheet[f'B{row}']
                    cell_c = worksheet[f'C{row}']
                    cell_b.font = Font(size=10)
                    cell_c.font = Font(size=10)
                    cell_b.alignment = Alignment(wrap_text=True, vertical='center')
                    cell_c.alignment = Alignment(wrap_text=True, vertical='center')
                    worksheet.row_dimensions[row].height = 80

                # 设置列宽（单位为字符宽度）
                worksheet.column_dimensions['A'].width = 22
                worksheet.column_dimensions['B'].width = 22
                worksheet.column_dimensions['C'].width = 22
                worksheet.column_dimensions['D'].width = 22
                worksheet.column_dimensions['E'].width = 22
        except Exception as e:
            print(f"Error writing to Excel: {str(e)}")
            raise

    def store_api_key(self, api_key: str) -> None:
        """存储API Key"""
        self.api_key = api_key
        if self.status_label:
            self.status_label.setText("已存储API Key")

    def generate_ai_report(self, num_objectives: int, current_achievement: Dict[str, float]) -> None:
        """生成AI分析报告"""
        if not self.api_key:
            raise ValueError("请先设置API Key")
        
        course_name = self.course_name_input.text()
        if not course_name:
            if self.status_label:
                self.status_label.setText("请先输入课程名称")
            return
        
        try:
            self.generate_improvement_report(current_achievement, course_name, num_objectives)
            if self.status_label:
                self.status_label.setText("AI分析报告已生成")
        except Exception as e:
            if self.status_label:
                self.status_label.setText("生成AI分析报告失败！")
            raise ValueError(f"生成AI分析报告失败: {str(e)}")